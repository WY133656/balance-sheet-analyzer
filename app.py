"""
资产负债表智能分析工具
纯 AI 驱动 — DeepSeek 直接理解报表内容
单文件架构 — 零模块 import，杜绝缓存问题
"""
import streamlit as st
import pandas as pd
import json
import io
from openai import OpenAI

# ====== 页面配置 ======
st.set_page_config(page_title="资产负债表分析", page_icon="📊", layout="wide")

# ====== 文件读取（内联） ======
def read_file(file) -> str:
    """读取 Excel 或 PDF，返回纯文本"""
    name = file.name.lower()
    if name.endswith(('.xlsx', '.xls')):
        return _read_excel(file)
    elif name.endswith('.pdf'):
        return _read_pdf(file)
    return ""

def _read_excel(file) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    parts = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        parts.append(f"\n===== 工作表: {sn} =====\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts)

def _read_pdf(file) -> str:
    import pdfplumber
    parts = []
    with pdfplumber.open(io.BytesIO(file.read())) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                parts.append(f"\n--- 第{i+1}页 ---\n")
                parts.append(text)
            tables = page.extract_tables()
            for t in tables:
                if t:
                    for row in t:
                        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                        if cells:
                            parts.append(" | ".join(cells))
    return "\n".join(parts)

# ====== DeepSeek 调用（内联） ======
def ask_deepseek(prompt: str, api_key: str, model: str = "deepseek-chat") -> str:
    """调用 DeepSeek API"""
    client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com/v1")
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": "你是一位资深财务分析师，专精资产负债表分析。请始终用中文回复，数据准确，逻辑清晰。"},
            {"role": "user", "content": prompt}
        ],
        temperature=0.2,
        max_tokens=2000
    )
    return response.choices[0].message.content

def extract_balance_sheet_data(raw_text: str, api_key: str, model: str) -> dict:
    """步骤1：让 AI 读懂资产负债表，提取结构化数据"""
    prompt = f"""请仔细阅读以下资产负债表内容，提取所有财务数据。

===== 报表原文 =====
{raw_text[:10000]}

===== 任务 =====
1. 判断数据单位（元/万元/亿元）
2. 提取所有资产类科目和金额
3. 提取所有负债类科目和金额
4. 提取所有所有者权益类科目和金额
5. 计算：总资产、总负债、所有者权益合计、资产负债率、流动比率（如果能找到流动资产和流动负债）
6. 如果报表有期初/期末两列，分别提取

===== 输出格式 =====
请严格输出 JSON，不要有额外文字：
{{
  "报告期间": "",
  "数据单位": "",
  "资产": {{"科目1": "金额1", "科目2": "金额2", ...}},
  "负债": {{"科目1": "金额1", ...}},
  "所有者权益": {{"科目1": "金额1", ...}},
  "关键指标": {{
    "总资产": "",
    "总负债": "",
    "所有者权益合计": "",
    "资产负债率": "",
    "流动比率": ""
  }},
  "值得关注的3个要点": ["", "", ""]
}}
"""
    reply = ask_deepseek(prompt, api_key, model)
    # 提取 JSON
    try:
        return json.loads(_extract_json(reply))
    except:
        return {"raw_reply": reply, "parse_error": True}

def generate_analysis(data: dict, api_key: str, model: str) -> str:
    """步骤2：基于提取的数据生成分析报告"""
    prompt = f"""请基于以下资产负债表数据，生成一份专业的财务分析报告。

===== 数据 =====
{json.dumps(data, ensure_ascii=False, indent=2)}

===== 分析要求 =====
请从以下角度分析（通俗易懂，每个角度 3-5 句话）：

1. **资产结构分析**：总资产规模、流动资产 vs 非流动资产占比、主要资产科目
2. **负债结构分析**：负债水平、流动负债 vs 非流动负债、主要负债科目
3. **偿债能力分析**：资产负债率是否合理、流动比率、短期偿债压力
4. **所有者权益分析**：净资产规模、资本结构是否稳健
5. **综合评估**：给出优秀/良好/关注/风险 的评级，并说明理由

请用中文输出，适合非财务专业人士阅读。"""
    return ask_deepseek(prompt, api_key, model)

def chat_about_data(data: dict, question: str, api_key: str, model: str) -> str:
    """步骤3：自由问答"""
    prompt = f"""资产负债表数据：{json.dumps(data, ensure_ascii=False)}

用户问题：{question}

请基于数据回答，用中文。如果数据不足以回答，请诚实说明。"""
    return ask_deepseek(prompt, api_key, model)

def _extract_json(text: str) -> str:
    """从 LLM 回复中提取 JSON"""
    import re
    m = re.search(r'```(?:json)?\s*([\s\S]*?)```', text)
    if m:
        return m.group(1).strip()
    start, end = text.find('{'), text.rfind('}')
    if start >= 0 and end > start:
        return text[start:end+1]
    return text

# ====== Session State 初始化 ======
for key, default in [
    ("raw_text", ""),
    ("file_name", ""),
    ("extracted_data", None),
    ("analysis_result", None),
    ("chat_history", []),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ====== 侧边栏 ======
with st.sidebar:
    st.title("⚙️ 配置")
    api_key = st.text_input(
        "DeepSeek API Key",
        type="password",
        placeholder="sk-...",
        help="在 platform.deepseek.com 注册获取"
    )
    model = st.selectbox("模型", ["deepseek-chat", "deepseek-reasoner"], index=0)

    st.markdown("---")
    st.markdown("### 📌 使用说明")
    st.markdown("""
    1. 填写 DeepSeek API Key
    2. 上传资产负债表（PDF/Excel）
    3. 点击「开始分析」
    4. 查看 AI 提取的数据和分析报告
    5. 在问答区自由提问
    """)
    st.markdown("---")
    st.markdown("### 💡 提示")
    st.markdown("- 支持上市公司年报 PDF")
    st.markdown("- 支持 Excel 格式报表")
    st.markdown("- DeepSeek 免费额度够用")
    st.markdown("- 数据仅在本次会话中使用")

# ====== 主页面 ======
st.title("📊 资产负债表智能分析")
st.markdown("上传资产负债表，DeepSeek 自动提取数据并生成专业分析报告")

# 文件上传
uploaded = st.file_uploader(
    "📁 上传资产负债表（PDF / Excel）",
    type=["pdf", "xlsx", "xls"],
    label_visibility="collapsed"
)

if uploaded:
    # 读取文件
    if st.session_state.file_name != uploaded.name:
        with st.spinner("正在读取文件..."):
            raw = read_file(uploaded)
            st.session_state.raw_text = raw
            st.session_state.file_name = uploaded.name
            st.session_state.extracted_data = None
            st.session_state.analysis_result = None
            st.session_state.chat_history = []

    st.success(f"✅ {uploaded.name} 已读取")

    # Tab 布局
    tab0, tab1, tab2, tab3 = st.tabs(["📄 原文预览", "🔢 数据提取", "📝 分析报告", "💬 智能问答"])

    # ==== Tab 0: 原文预览 ====
    with tab0:
        st.caption("以下是文件中读取到的原始内容")
        st.text_area("原始内容", st.session_state.raw_text[:6000], height=400)

    # ==== Tab 1: 数据提取 ====
    with tab1:
        st.subheader("🔢 AI 提取的资产负债表数据")

        if st.button("🔄 开始提取数据", type="primary"):
            if not api_key:
                st.error("请先在侧边栏填写 DeepSeek API Key")
            elif not st.session_state.raw_text:
                st.error("未能读取到文件内容")
            else:
                with st.spinner("DeepSeek 正在理解资产负债表..."):
                    data = extract_balance_sheet_data(
                        st.session_state.raw_text, api_key, model
                    )
                    st.session_state.extracted_data = data

        if st.session_state.extracted_data:
            d = st.session_state.extracted_data
            if d.get("parse_error"):
                st.warning("JSON 解析异常，显示原始回复")
                st.text(d.get("raw_reply", ""))
            else:
                # 三列展示
                c1, c2, c3 = st.columns(3)
                with c1:
                    st.metric("🏢 报告期间", d.get("报告期间", "-"))
                with c2:
                    st.metric("📐 数据单位", d.get("数据单位", "-"))
                with c3:
                    kpi = d.get("关键指标", {})
                    st.metric("📊 资产负债率", kpi.get("资产负债率", "-"))

                # 资产表
                assets = d.get("资产", {})
                if assets:
                    st.markdown("### 🟢 资产")
                    st.dataframe(
                        pd.DataFrame({"科目": assets.keys(), "金额": assets.values()}),
                        use_container_width=True
                    )

                # 负债表
                liab = d.get("负债", {})
                if liab:
                    st.markdown("### 🔴 负债")
                    st.dataframe(
                        pd.DataFrame({"科目": liab.keys(), "金额": liab.values()}),
                        use_container_width=True
                    )

                # 所有者权益表
                equity = d.get("所有者权益", {})
                if equity:
                    st.markdown("### 🔵 所有者权益")
                    st.dataframe(
                        pd.DataFrame({"科目": equity.keys(), "金额": equity.values()}),
                        use_container_width=True
                    )

                # 关键指标
                if kpi:
                    st.markdown("### 📊 关键指标")
                    kpi_cols = st.columns(len(kpi))
                    for i, (k, v) in enumerate(kpi.items()):
                        with kpi_cols[i]:
                            st.metric(k, str(v) if v else "-")

                # 要点
                points = d.get("值得关注的3个要点", [])
                if points:
                    st.markdown("### ⚡ 要点")
                    for p in points:
                        if p:
                            st.markdown(f"- {p}")

    # ==== Tab 2: 分析报告 ====
    with tab2:
        st.subheader("📝 AI 财务分析报告")

        if st.button("📝 生成分析报告", type="primary"):
            if not api_key:
                st.error("请先在侧边栏填写 DeepSeek API Key")
            elif not st.session_state.extracted_data:
                st.warning("请先在「数据提取」中点击提取数据")
            else:
                with st.spinner("DeepSeek 正在撰写分析报告..."):
                    report = generate_analysis(
                        st.session_state.extracted_data, api_key, model
                    )
                    st.session_state.analysis_result = report

        if st.session_state.analysis_result:
            st.markdown("---")
            st.markdown(st.session_state.analysis_result)

    # ==== Tab 3: 智能问答 ====
    with tab3:
        st.subheader("💬 基于数据的智能问答")
        st.caption("针对刚才提取的资产负债表数据自由提问")

        for msg in st.session_state.chat_history:
            with st.chat_message(msg["role"]):
                prefix = "👤 你" if msg["role"] == "user" else "🤖 AI"
                st.markdown(f"**{prefix}**：{msg['content']}")

        q = st.chat_input("例如：这家公司的负债结构是否健康？")
        if q:
            if not api_key:
                st.error("请先填写 API Key")
            elif not st.session_state.extracted_data:
                st.warning("请先提取数据")
            else:
                st.session_state.chat_history.append({"role": "user", "content": q})
                with st.spinner("AI 思考中..."):
                    answer = chat_about_data(
                        st.session_state.extracted_data, q, api_key, model
                    )
                st.session_state.chat_history.append({"role": "assistant", "content": answer})
                st.rerun()

else:
    # 欢迎页
    st.markdown("---")
    st.markdown("""
    ### 👋 欢迎使用资产负债表智能分析工具

    **三步获得专业分析：**

    | 步骤 | 操作 |
    |------|------|
    | 1️⃣ | 填写 DeepSeek API Key（侧边栏） |
    | 2️⃣ | 上传资产负债表 PDF 或 Excel |
    | 3️⃣ | 点击「开始提取」→「生成报告」 |

    ---

    ### 🔍 支持分析的内容

    - ✅ **资产结构** — 流动资产、固定资产、无形资产等占比分析
    - ✅ **负债结构** — 流动负债、长期负债、负债水平评估
    - ✅ **偿债能力** — 资产负债率、流动比率、速动比率
    - ✅ **所有者权益** — 净资产规模、资本结构稳健性
    - ✅ **智能问答** — 基于数据自由提问

    ---

    🔧 **技术栈**：Python · Streamlit · DeepSeek · pdfplumber · openpyxl

    💡 [注册 DeepSeek](https://platform.deepseek.com) 获取 API Key（新用户免费额度）
    """)

st.markdown("---")
st.caption("© 2025 资产负债表智能分析 | AI 作品集项目")
